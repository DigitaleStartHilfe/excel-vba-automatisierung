"""
Excel + Python Automatisierung
Projekt: Excel-Tabellen-Vergleicher (Demo 6)
Erstellt: August 2026
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import os

# 1. Testdaten erstellen (falls nicht vorhanden)
if not os.path.exists("soll.xlsx"):
    soll_daten = {
        "Name": ["Anna", "Ben", "Clara", "David", "Eva"],
        "Abteilung": ["Vertrieb", "IT", "Marketing", "Vertrieb", "IT"],
        "Gehalt": [4200, 5100, 3800, 4900, 5600],
        "Bonus": [420, 510, 380, 490, 560]
    }
    pd.DataFrame(soll_daten).to_excel("soll.xlsx", index=False)
    print("📄 Soll-Datei erstellt")

if not os.path.exists("ist.xlsx"):
    ist_daten = {
        "Name": ["Anna", "Ben", "Clara", "David", "Eva"],
        "Abteilung": ["Vertrieb", "IT", "Marketing", "Vertrieb", "IT"],
        "Gehalt": [4200, 5300, 3800, 4900, 5600],  # Abweichung bei Ben
        "Bonus": [420, 530, 380, 490, 560]
    }
    pd.DataFrame(ist_daten).to_excel("ist.xlsx", index=False)
    print("📄 Ist-Datei erstellt")

# 2. Vergleichen
df_soll = pd.read_excel("soll.xlsx")
df_ist = pd.read_excel("ist.xlsx")

with pd.ExcelWriter("vergleich_ergebnis.xlsx", engine="openpyxl") as writer:
    df_ist.to_excel(writer, sheet_name="Vergleich", index=False)
    wb = writer.book
    ws = writer.sheets["Vergleich"]
    gruen = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rot = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for idx in range(2, len(df_ist) + 2):
        if df_soll.iloc[idx-2].equals(df_ist.iloc[idx-2]):
            for col in range(1, len(df_ist.columns) + 1):
                ws.cell(row=idx, column=col).fill = gruen
        else:
            for col in range(1, len(df_ist.columns) + 1):
                ws.cell(row=idx, column=col).fill = rot

print("✅ Vergleich abgeschlossen! Siehe 'vergleich_ergebnis.xlsx'")
